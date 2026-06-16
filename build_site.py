#!/usr/bin/env python3
"""
build_site.py — 读取 arena_leaderboard.xlsx 生成 Tab 式 HTML 网页
=================================================================
与 export_arena_excel.py 配套使用：先运行 export_arena_excel.py 生成 xlsx，
再运行本脚本生成 index.html 用于 GitHub Pages 展示。

布局与 Excel 一致，列由 xlsx 动态驱动（Excel 加列，网页自动继承）：
  - 汇总总览页：各维度 Top-10 竖向柱状图（国别配色）
  - Elo 榜：排名 / 模型名 / 厂商(产品 logo) / 许可 / 分数(数值+置信区间) / 投票数
            + price / context 等原始列 1:1 复刻
  - Agent 榜：行为指标列（含中文说明 tooltip）；"净提升"按排名绿→红着色
  - 前三名金/银/铜高亮，许可类型仅文字着色
"""
import re
import io
import sys
import json
import base64
from html.parser import HTMLParser
from urllib.parse import urljoin
from datetime import datetime, timezone, timedelta
from pathlib import Path

import requests
from openpyxl import load_workbook

try:                       # Pillow 用于把官网图标统一缩放为 64px PNG；缺失时降级
    from PIL import Image
    _HAS_PIL = True
except ImportError:
    _HAS_PIL = False


# ── 各榜单介绍（译自 arena.ai 官方文案，按 Excel sheet 名索引） ──
CATEGORY_DESC = {
    "📝 文本-对话 (Text)": "在数学、编程、创意写作及其他开放式领域，纵览各类 AI 模型在文本到文本任务上的综合排名。",
    "💻 代码 (Code)": "纵览各类 AI 模型在前端 Web 开发任务上的综合排名，涵盖需要多步推理与工具调用的 Agentic 编程工作流。",
    "👁️ 视觉理解 (Vision)": "纵览具备视觉输入推理能力的多模态 AI 模型的综合排名。",
    "📄 文档处理 (Document)": "纵览各类 AI 模型在文档分析与长文本推理上的综合排名。",
    "🔍 搜索增强 (Search)": "纵览集成网络搜索能力的大语言模型的综合排名。",
    "🤖 智能体 (Agent)": "纵览各 AI 模型在真实 Agentic 任务中编排工具能力的动态排名，依据工具可靠性、任务完成度、可操控性等行为信号。",
    "🎨 文生图 (Text-to-Image)": "纵览各类文生图 AI 模型的综合排名。",
    "✂️ 图片编辑 (Image-Edit)": "纵览各类图像编辑 AI 模型的综合排名。",
    "🎬 文生视频 (Text-to-Video)": "纵览各类文生视频 AI 模型的综合排名。",
    "📹 图生视频 (Image-to-Video)": "纵览各类图生视频 AI 模型的综合排名。",
    "🎞️ 视频编辑 (Video-Edit)": "纵览各类视频编辑 AI 模型的综合排名。",
}


# Agent 榜各行为指标的中文释义（译自官方方法论页 arena.ai/blog/agent-arena-methodology），
# 用于在列头显示 (i) 图标的悬停提示。键需与 export 生成的列名完全一致。
METRIC_TOOLTIPS = {
    "净提升 (Net Improvement)": "净提升：把智能体视为多组件系统，通过随机化各组件选择构成随机对照试验，用因果推断估计“选用该模型作为协调器”带来的因果处理效应（即整体表现的净提升）；榜单总分由下方各信号聚合而成，误差为 95% 置信区间。",
    "确认完成 (Confirmed Success)": "确认完成：用户通过 Arena 界面的“赞同/不赞同”按钮标记任务成败，以该任务轨迹最终的认可与否为准（一次会话可能包含多个任务）。",
    "好评/差评 (Praise vs Complaint)": "好评 vs 差评：统计任务中明确的口头好评（如“看起来很棒”“正是我要的”）与差评（如“这是坏的”“你完全理解错了”）；好评多于差评则该任务记为成功。",
    "可操控性 (Steerability)": "可操控性：用户中途纠正（如“不，改成做 X”“你读错文件了”）时，智能体执行修正的能力。用户接受修正记为成功，拒绝或放弃则记为失败——衡量出错后能否被快速纠偏。",
    "命令恢复 (Bash Recovery)": "命令恢复：从 bash 报错中恢复所需的回合数。当智能体因自身（而非环境）原因执行出错命令时开始计数，统计到下一条不报错命令之间的 bash 调用次数；若放弃则额外扣分。",
    "工具幻觉 (Tool Hallucination)": "工具幻觉：智能体引用了并不存在的工具。会惩罚凭空捏造的工具名、语法错误产生的无效名、以及思维链文本误入工具字段等；只要调用了不存在的工具，该任务即记为失败。",
    "会话数 (Sessions)": "会话数：用于评估该模型的真实 Agent Mode 交互会话数量（一次会话可能包含多个任务）；数量越多，统计估计越可靠。",
}


def _heat(t: float) -> str:
    """t∈[0,1]：0→红, 0.5→黄, 1→绿（Excel 经典三色阶）。"""
    t = max(0.0, min(1.0, t))
    if t <= 0.5:
        f = t / 0.5
        a, b = (248, 105, 107), (255, 235, 132)
    else:
        f = (t - 0.5) / 0.5
        a, b = (255, 235, 132), (99, 190, 123)
    return "rgb({},{},{})".format(
        round(a[0] + (b[0] - a[0]) * f),
        round(a[1] + (b[1] - a[1]) * f),
        round(a[2] + (b[2] - a[2]) * f),
    )


# 厂商 → 其大模型【产品】域名（用于取产品 logo，而非母公司）。
# 取该域名的 favicon 作图标；未命中则回退首字母方块。
VENDOR_DOMAIN = {
    "openai": "openai.com", "anthropic": "claude.ai",
    "google": "gemini.google.com", "google deepmind": "deepmind.google",
    "deepmind": "deepmind.google", "meta": "llama.com", "xai": "x.ai",
    "deepseek": "deepseek.com", "alibaba": "qwen.ai", "alibaba-ath": "qwen.ai",
    "qwen": "qwen.ai", "z.ai": "z.ai", "zhipu": "z.ai", "01.ai": "01.ai",
    "moonshot": "kimi.com", "minimax": "minimax.io", "mistral": "mistral.ai",
    "perplexity": "perplexity.ai", "microsoft": "microsoft.com",
    "amazon": "aws.amazon.com", "nvidia": "nvidia.com", "cohere": "cohere.com",
    "ai21": "ai21.com", "baidu": "yiyan.baidu.com", "xiaomi": "mi.com",
    "tencent": "hunyuan.tencent.com", "tencent hunyuan": "hunyuan.tencent.com",
    "bytedance": "dreamina.com", "kuaishou": "klingai.com", "kling": "klingai.com",
    "klingai": "klingai.com", "midjourney": "midjourney.com",
    "ideogram": "ideogram.ai", "adobe": "firefly.adobe.com", "runway": "runwayml.com",
    "luma": "lumalabs.ai", "luma ai": "lumalabs.ai", "genmo": "genmo.ai",
    "genmo ai": "genmo.ai", "stability ai": "stability.ai",
    "black forest labs": "bfl.ai", "vidu": "vidu.com", "lightricks": "lightricks.com",
    "reve": "reve.art", "recraft": "recraft.ai", "pika": "pika.art",
    "hailuo": "hailuoai.com", "pixverse": "pixverse.ai", "reka": "reka.ai",
    "databricks": "databricks.com", "ibm": "ibm.com", "snowflake": "snowflake.com",
    "hidream": "hidream.ai", "pruna": "pruna.ai",
    "huggingface": "huggingface.co", "meituan": "meituan.com", "krea": "krea.ai",
    "ai2": "allenai.org", "allenai": "allenai.org", "lmsys": "lmsys.org",
    "upstage": "upstage.ai", "inception": "inceptionlabs.ai",
}


def _vendor_domain(vendor: str):
    v = (vendor or "").strip().lower()
    if v in VENDOR_DOMAIN:
        return VENDOR_DOMAIN[v]
    v2 = v.replace(".ai", "").replace(" ai", "").strip()   # 容错：去掉常见后缀再试
    return VENDOR_DOMAIN.get(v2)


# 模型名关键词 → (厂商显示名, 产品域名)。arena 厂商列很稀疏（多数行不带厂商徽标），
# 据模型名补出厂商与产品 logo，避免图标缺失。按特异性从前往后匹配。
MODEL_BRAND = [
    ("claude", "Anthropic", "claude.ai"),
    ("gpt-image", "OpenAI", "openai.com"), ("chatgpt", "OpenAI", "openai.com"),
    ("dall-e", "OpenAI", "openai.com"), ("dalle", "OpenAI", "openai.com"),
    ("sora", "OpenAI", "openai.com"), ("gpt", "OpenAI", "openai.com"),
    ("nano-banana", "Google", "gemini.google.com"), ("gemini", "Google", "gemini.google.com"),
    ("veo", "Google", "gemini.google.com"), ("imagen", "Google", "gemini.google.com"),
    ("gemma", "Google", "gemini.google.com"),
    ("grok", "xAI", "x.ai"), ("glm", "Z.ai", "z.ai"),
    ("qwen", "Qwen", "qwen.ai"), ("qwq", "Qwen", "qwen.ai"), ("wan", "Alibaba", "qwen.ai"),
    ("deepseek", "DeepSeek", "deepseek.com"), ("llama", "Meta", "llama.com"),
    ("mixtral", "Mistral", "mistral.ai"), ("codestral", "Mistral", "mistral.ai"),
    ("magistral", "Mistral", "mistral.ai"), ("mistral", "Mistral", "mistral.ai"),
    ("kimi", "Moonshot", "kimi.com"),
    ("minimax", "MiniMax", "minimax.io"), ("hailuo", "MiniMax", "minimax.io"),
    ("abab", "MiniMax", "minimax.io"),
    ("dreamina", "Bytedance", "dreamina.com"), ("seedance", "Bytedance", "dreamina.com"),
    ("seedream", "Bytedance", "dreamina.com"), ("doubao", "Bytedance", "dreamina.com"),
    ("kling", "Kuaishou", "klingai.com"), ("ernie", "Baidu", "yiyan.baidu.com"),
    ("hunyuan", "Tencent", "hunyuan.tencent.com"), ("flux", "Black Forest Labs", "bfl.ai"),
    ("reve", "Reve", "reve.art"), ("ideogram", "Ideogram", "ideogram.ai"),
    ("midjourney", "Midjourney", "midjourney.com"), ("recraft", "Recraft", "recraft.ai"),
    ("vidu", "Vidu", "vidu.com"), ("pika", "Pika", "pika.art"),
    ("runway", "Runway", "runwayml.com"), ("gen-4", "Runway", "runwayml.com"),
    ("aleph", "Runway", "runwayml.com"), ("luma", "Luma AI", "lumalabs.ai"),
    ("firefly", "Adobe", "firefly.adobe.com"), ("mai-", "Microsoft", "microsoft.com"),
    ("phi-", "Microsoft", "microsoft.com"), ("command", "Cohere", "cohere.com"),
    ("pixverse", "Pixverse", "pixverse.ai"), ("stable", "Stability AI", "stability.ai"),
    ("sdxl", "Stability AI", "stability.ai"),
    ("yi-", "01.AI", "01.ai"), ("reka", "Reka", "reka.ai"),
    ("jamba", "AI21", "ai21.com"), ("mercury", "Inception", "inceptionlabs.ai"),
    ("longcat", "Meituan", "meituan.com"), ("krea", "Krea", "krea.ai"),
    ("ppl-sonar", "Perplexity", "perplexity.ai"), ("sonar", "Perplexity", "perplexity.ai"),
    ("solar", "Upstage", "upstage.ai"), ("kat-", "Kwai", "kuaishou.com"),
]


def _brand_from_model(model: str):
    m = (model or "").strip().lower()
    if m:
        for kw, name, dom in MODEL_BRAND:
            if kw in m:
                return name, dom
    return "", None


_GENERIC_VENDOR = {"ai", "labs", "lab", "group", "inc", "team", "research",
                   "technologies", "technology", "co", "ltd"}


def _brand(vendor: str, model: str):
    """该行的（厂商显示名, 产品域名）：优先用抓到的厂商；为空或为泛词（如 arena 只渲染出
    'AI'/'Labs'/'Group' 等多词厂商末尾词）时，改用模型名推断品牌，避免误显示与图标缺失。"""
    v = (vendor or "").strip()
    if v and re.sub(r"[^a-z0-9]", "", v.lower()) not in _GENERIC_VENDOR:
        return v, (_vendor_domain(v) or _brand_from_model(model)[1])
    bn, bd = _brand_from_model(model)         # 泛词/缺失 → 模型名推断
    return (bn or v), (bd or _vendor_domain(v))


# ── 厂商图标：构建时全部下载并内嵌，运行时零网络请求（保证各环境/离线渲染一致）──
# 取图优先级：① arena.ai 自带的内联品牌 SVG（矢量、最贴合官网）→ ② 厂商官网 favicon
# → ③ favicon 服务；三者都在【构建时】内嵌为 data URI。都失败才用首字母方块。
# 不再有任何运行时第三方请求（此前用 Google favicon 服务实时加载，在屏蔽 Google 的
# 网络环境下会全部失败退化成字母——这是跨环境不一致的根因）。
# 结果缓存到 data/logos_cache.json（提速 + 容灾），并随每日数据提交回仓库。
_LOGO_UA = {"User-Agent": "Mozilla/5.0 (compatible; arena-leaderboard-logo/1.0)"}
_LOGO_TIMEOUT = 8
_RESOLVED_LOGOS: dict = {}        # key(域名/规范名) -> data URI（""=已尝试失败）
_LOGO_CACHE_PATH = Path(__file__).parent / "data" / "logos_cache.json"
_LOGO_REFRESH = False
_logo_session = None

# arena.ai 提供品牌 SVG 的厂商（标题命名）→ 与本项目厂商显示名的别名对齐
_ARENA_BOARDS = ["text", "code", "agent", "text-to-image", "text-to-video",
                 "image-to-video", "vision"]
_ARENA_ALIAS = {            # 本项目名(规范化) -> arena 标题(规范化)
    "blackforestlabs": "flux", "kuaishou": "kwai", "kling": "kwai",
    "stabilityai": "stability", "01ai": "01ai", "stepfun": "stepfun",
    "kwaikat": "kwai",
}
_arena_svgs = None          # 规范化标题 -> svg data URI（懒加载）


def _norm(s: str) -> str:
    """厂商名规范化：仅留小写字母数字（用于跨命名匹配）。"""
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _session():
    global _logo_session
    if _logo_session is None:
        _logo_session = requests.Session()
        _logo_session.headers.update(_LOGO_UA)
    return _logo_session


def _load_arena_svgs() -> dict:
    """抓取 arena.ai 各榜单页，提取其内联品牌 SVG：{规范化标题: svg data URI}。"""
    global _arena_svgs
    if _arena_svgs is not None:
        return _arena_svgs
    _arena_svgs = {}
    for board in _ARENA_BOARDS:
        try:
            html = _session().get(f"https://arena.ai/leaderboard/{board}",
                                  timeout=_LOGO_TIMEOUT).text
        except Exception:         # noqa: BLE001 — 某榜单取不到不影响其它
            continue
        for m in re.finditer(r"<svg\b[^>]*>.*?</svg>", html, re.S):
            block = m.group(0)
            tm = re.search(r"<title>(.*?)</title>", block, re.S)
            if not tm:
                continue
            title = re.sub(r"\s+", " ", tm.group(1)).strip()
            key = _norm(title)
            if not key or len(title) > 24 or key in _arena_svgs:
                continue          # 跳过页面大标题等非厂商项
            svg = re.sub(r'\sclass="[^"]*"', "", block)       # 仅去掉无效 class，其余原样保留
            # 注：不可删除内部元素的 width/height（如 Bytedance 用 <rect> 定义形状，删后会坍缩）
            _arena_svgs[key] = "data:image/svg+xml;base64," + \
                base64.b64encode(svg.encode()).decode()
    print(f"  arena 品牌 SVG: 收集到 {len(_arena_svgs)} 个")
    return _arena_svgs


def _resolve_from_arena(name: str):
    """① arena.ai 自带品牌 SVG（按厂商名匹配，含别名）。"""
    svgs = _load_arena_svgs()
    n = _norm(name)
    if n in svgs:
        return svgs[n]
    if _ARENA_ALIAS.get(n) in svgs:
        return svgs[_ARENA_ALIAS[n]]
    return None


class _IconLinkParser(HTMLParser):
    """从 HTML 抽取 <link rel=...icon...> 候选。"""
    def __init__(self):
        super().__init__()
        self.icons = []

    def handle_starttag(self, tag, attrs):
        if tag != "link":
            return
        d = dict(attrs)
        rel = (d.get("rel") or "").lower()
        if "icon" in rel and d.get("href"):
            self.icons.append((rel, d["href"]))


def _normalize_icon(raw: bytes) -> str:
    """图标字节 → data URI。SVG 原样内嵌；位图统一缩放为 ≤64px PNG。"""
    head = raw[:256].lstrip().lower()
    if head[:5] == b"<?xml" or b"<svg" in head:
        return "data:image/svg+xml;base64," + base64.b64encode(raw).decode()
    if _HAS_PIL:
        im = Image.open(io.BytesIO(raw)).convert("RGBA")
        if max(im.size) > 64:
            im.thumbnail((64, 64), Image.LANCZOS)
        buf = io.BytesIO()
        im.save(buf, "PNG")
        raw = buf.getvalue()
    elif len(raw) > 25600:        # 无 Pillow 又过大：放弃，交由上层回退
        raise ValueError("icon too large without Pillow")
    return "data:image/png;base64," + base64.b64encode(raw).decode()


def _resolve_from_official(domain: str):
    """直抓官网：解析首页 icon link（取 apple-touch/png/svg 优先），下载并归一。"""
    base = f"https://{domain}/"
    r = _session().get(base, timeout=_LOGO_TIMEOUT)
    if r.status_code != 200:
        return None
    p = _IconLinkParser()
    p.feed(r.text)

    def score(item):
        rel, href = item
        h = href.lower()
        return (("apple-touch" in rel) * 3 + h.endswith(".png") * 2 + h.endswith(".svg"))

    cands = sorted(p.icons, key=score, reverse=True)
    urls = [urljoin(base, h) for _, h in cands] + [urljoin(base, "/favicon.ico")]
    for u in urls[:4]:
        try:
            ir = _session().get(u, timeout=_LOGO_TIMEOUT)
            if ir.status_code == 200 and ir.content:
                return _normalize_icon(ir.content)
        except Exception:        # noqa: BLE001 — 单个候选失败试下一个
            continue
    return None


def _resolve_from_service(domain: str):
    """回退：用 Google favicon 服务取图并内嵌（仍自托管，无运行时第三方请求）。"""
    u = f"https://www.google.com/s2/favicons?domain={domain}&sz=64"
    ir = _session().get(u, timeout=_LOGO_TIMEOUT)
    if ir.status_code == 200 and ir.content:
        return _normalize_icon(ir.content)
    return None


def _logo_key(name: str, domain: str) -> str:
    """该厂商的稳定缓存键 / CSS 类基名：有域名用域名，否则用规范化名。"""
    return domain or (f"n-{_norm(name)}" if name else "")


def _get_logo(name: str, domain: str) -> str:
    """返回内嵌图标 data URI（arena SVG→官网→服务），全部失败返回 ""（上层用字母）。"""
    key = _logo_key(name, domain)
    if not key:
        return ""
    if not _LOGO_REFRESH and key in _RESOLVED_LOGOS:
        return _RESOLVED_LOGOS[key]
    uri = ""
    chain = [("arena 官网图标", lambda: _resolve_from_arena(name)),
             ("厂商官网", lambda: _resolve_from_official(domain) if domain else None),
             ("favicon 服务", lambda: _resolve_from_service(domain) if domain else None)]
    for label, fn in chain:
        try:
            got = fn()
        except Exception:        # noqa: BLE001 — 网络/解析异常视为该路失败
            got = None
        if got:
            uri = got
            print(f"  logo: {name or domain} ← {label}")
            break
    if not uri:
        print(f"  logo: {name or domain} 未取到，使用字母")
    _RESOLVED_LOGOS[key] = uri
    return uri


def load_logo_cache():
    if _LOGO_CACHE_PATH.exists():
        try:
            _RESOLVED_LOGOS.update(json.loads(_LOGO_CACHE_PATH.read_text(encoding="utf-8")))
        except (ValueError, OSError):
            pass


def save_logo_cache():
    try:
        _LOGO_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        # 只缓存成功项，失败项留空以便下次重试
        data = {k: v for k, v in _RESOLVED_LOGOS.items() if v}
        _LOGO_CACHE_PATH.write_text(
            json.dumps(data, ensure_ascii=False, indent=0), encoding="utf-8")
    except OSError:
        pass


def _logo_slug(key: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", key.lower()).strip("-")


def logo_css() -> str:
    """把已解析的图标按 key 各生成一条 CSS 背景规则（每个图标只出现一次，避免逐行重复内嵌）。"""
    rules = [".vlogo.img{background-size:cover;background-position:center;background-repeat:no-repeat}",
             ".vlogo.img::before{content:none}"]
    for key, uri in sorted(_RESOLVED_LOGOS.items()):
        if uri:
            rules.append(f'.l-{_logo_slug(key)}{{background-image:url({uri})}}')
    return "\n".join(rules)


def _vendor_cell_html(vendor: str, model: str) -> str:
    """厂商单元格：产品 logo + 厂商名（厂商列稀疏时据模型名补全，避免图标缺失）。"""
    name, domain = _brand(vendor, model)
    label = name or (vendor or "").strip()
    letter = _esc(((label or model or "·")[:1]).upper())
    uri = _get_logo(name, domain) if (name or domain) else ""
    if uri:                                  # 构建时已内嵌：用 CSS 类引用（每图标仅定义一次）
        logo = f'<span class="vlogo img l-{_logo_slug(_logo_key(name, domain))}" data-l="{letter}"></span>'
    else:                                    # 取不到：首字母方块（不依赖任何运行时网络）
        logo = f'<span class="vlogo" data-l="{letter}"></span>'
    return f'{logo}<span class="vname">{_esc(name)}</span>'


def _logo_only(vendor: str, model: str) -> str:
    """只返回产品 logo 图标（不带厂商名），用于汇总柱状图等紧凑场景。"""
    name, domain = _brand(vendor, model)
    label = name or (vendor or "").strip()
    letter = _esc(((label or model or "·")[:1]).upper())
    uri = _get_logo(name, domain) if (name or domain) else ""
    cls = f" img l-{_logo_slug(_logo_key(name, domain))}" if uri else ""
    return f'<span class="vlogo{cls}" data-l="{letter}" title="{_esc(name or label)}"></span>'


_SPREAD_RE = re.compile(r"(\d+)\s*[-–—~]\s*(\d+)")


def _fmt_rank_spread(text: str) -> str:
    """排名区间：把两个数之间的连接符统一成波浪线，如 '1-2'/'1 2' → '1~2'。"""
    t = (text or "").strip()
    if not t:
        return t
    m = _SPREAD_RE.search(t)
    if not m:
        m = re.search(r"(\d+)\s+(\d+)", t)        # 仅空格分隔
    return f"{m.group(1)}~{m.group(2)}" if m else t


def extract_sheets(xlsx_path: str) -> dict:
    """读取 xlsx 所有 Sheet，返回 {sheet_name: [行列表]}"""
    wb = load_workbook(xlsx_path, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
            rows.append([str(v) if v is not None else "" for v in row])
        sheets[name] = rows
    wb.close()
    return sheets


def _esc(text: str) -> str:
    """HTML 转义"""
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _category_top(rows: list, n: int = 10):
    """从一个类别 sheet 取前 n 名的 (模型, 厂商, 分数)。无 Elo 分数（如 Agent）返回 None。"""
    headers = rows[3] if len(rows) > 3 else []
    data = rows[4:] if len(rows) > 4 else []
    i_model = _find(headers, "模型")
    i_vendor = _find(headers, "厂商")
    i_score = _find(headers, "elo", "分数")
    if i_score < 0:
        return None
    out = []
    for row in data:
        sc = _num(row[i_score]) if i_score < len(row) else None
        if sc is None:
            continue
        out.append((row[i_model] if i_model < len(row) else "",
                    row[i_vendor] if i_vendor < len(row) else "", sc))
        if len(out) >= n:
            break
    return out


def render_summary_tab(category_sheets: list, now_str: str) -> str:
    """汇总总览：每个维度（除智能体）一行 Top-10 竖向柱状图，左高右低，柱内标分数，
    颜色仅区分国内/国外，参考 Artificial Analysis 风格。"""
    html = '<div class="panel-header summary-head">'
    html += ('<div><h2 class="panel-title">📊 全维度 Top 10 总览</h2>'
             '<p class="panel-subtitle">各能力维度（除智能体）前十名模型对比 · '
             '柱高反映相对分数 · 柱色区分国别</p></div>')
    html += (f'<div class="sb-updated">数据更新<b>{_esc(now_str)}</b>'
             f'<span>北京时间 · 来源 arena.ai</span></div>')
    html += '</div>'

    # 国别图例（只用国内/国外两色；未分类极少出现，灰色作静默兜底不入图例）
    html += ('<div class="sb-legend">'
             '<span class="lg-item"><span class="lg lg-fg"></span>国外模型</span>'
             '<span class="lg-item"><span class="lg lg-cn"></span>国内模型</span>'
             '</div>')

    for sname, rows in category_sheets:
        top = _category_top(rows, 10)
        if not top:                       # Agent 等无 Elo 分数 → 跳过
            continue
        title = sname.split("(")[0].strip() if "(" in sname else sname
        scores = [s for _, _, s in top]
        lo, hi = min(scores), max(scores)
        rng = (hi - lo) or 1.0

        html += '<div class="summary-board">'
        html += (f'<div class="sb-head"><h3 class="sb-title">{_esc(title)}</h3>'
                 f'<span class="sb-meta">Top {len(top)} · Elo {int(lo)}–{int(hi)}</span></div>')
        html += '<div class="sb-bars">'
        for rank, (model, vendor, sc) in enumerate(top, 1):
            region = _classify_region(vendor, model)
            h = 44 + (sc - lo) / rng * 56          # 柱高映射到 44%~100%，保证最矮也放得下分数
            disp = str(int(sc)) if sc == int(sc) else f"{sc:g}"
            html += (f'<div class="sb-bar {region}">'
                     f'<div class="sb-col"><div class="sb-fill" style="height:{h:.1f}%">'
                     f'<span class="sb-score">{_esc(disp)}</span></div></div>'
                     f'<div class="sb-logo">{_logo_only(vendor, model)}</div>'
                     f'<div class="sb-name" title="{_esc(model)}">{_esc(model)}</div>'
                     f'</div>')
        html += '</div></div>'

    return html


# ── 模型国别分类（柱色：cn 国内红 / fg 国外蓝 / unknown 未分类灰）──
# 数据外置在 data/regions.json，便于每日榜单更新时补充新厂商而无需改代码。
# 内置默认仅作兜底（文件缺失时仍可用）。判定以厂商组织名为准（优先于模型名前缀，
# 避免 LLaVA 这类"微调他人基座、名字含 qwen"被误判）。两表都未命中 → unknown 并告警。
_REGION_DEFAULT = {
    "cn": ["Alibaba", "Alibaba-ATH", "Qwen", "Baidu", "Bytedance", "DeepSeek",
           "Z.ai", "Zhipu", "Tsinghua", "Moonshot", "MiniMax", "Tencent",
           "Kuaishou", "Kling", "KlingAI", "Kwai", "KwaiKAT", "KAT", "01.AI",
           "Xiaomi", "Meituan", "Vidu", "Shengshu", "Pixverse", "HiDream",
           "StepFun", "OpenBMB", "InternLM", "OpenGVLab", "RWKV", "Ant",
           "InclusionAI", "Ling", "Skywork", "Baichuan", "SenseTime", "iFlytek"],
    "overseas": ["Anthropic", "OpenAI", "Google", "Meta", "xAI", "Mistral",
                 "Cohere", "Microsoft", "Amazon", "Nvidia", "AI21", "Perplexity",
                 "Reka", "Databricks", "IBM", "Snowflake", "Adobe", "Runway",
                 "Luma", "Luma AI", "Genmo", "Genmo AI", "Stability AI",
                 "Black Forest Labs", "Ideogram", "Midjourney", "Recraft",
                 "Reve", "Pika", "Krea", "Lightricks", "Upstage", "Ai2",
                 "AllenAI", "Inception", "Kandinsky", "TII", "Diffbot", "Pruna",
                 "NexusFlow", "NousResearch", "OpenAssistant", "OpenChat",
                 "MosaicML", "LMSYS", "Berkeley", "Princeton", "Stanford", "UW",
                 "Computations", "Intellect", "HuggingFace", "LLaVA"],
    "cn_model_prefixes": ["qwen", "qwq", "glm", "chatglm", "deepseek", "kimi",
                          "minimax", "hailuo", "ernie", "hunyuan", "doubao",
                          "dreamina", "seedance", "seedream", "kling", "kat-",
                          "vidu", "pixverse", "hidream", "longcat", "internlm",
                          "internvl", "skywork", "ling-", "ring-", "step-", "wan-",
                          "yi-", "baichuan", "minicpm", "rwkv"],
    "overseas_model_prefixes": ["trinity", "dolphin", "tulu", "starling",
                                "vicuna", "guanaco", "alpaca", "openhermes",
                                "oasst", "mpt", "athene", "falcon",
                                "stripedhyena", "lucid", "kandinsky"],
}
_REGION = None              # 规范化后的分类集合（懒加载）
_UNCLASSIFIED = {}          # 本次构建未分类的厂商：规范名 -> (显示名, 样例模型)


def _load_regions() -> dict:
    """加载国别分类表：data/regions.json 优先，缺失/出错回退内置默认。规范化键以便匹配。"""
    global _REGION
    if _REGION is not None:
        return _REGION
    raw = dict(_REGION_DEFAULT)
    fp = Path(__file__).parent / "data" / "regions.json"
    if fp.exists():
        try:
            ext = json.loads(fp.read_text(encoding="utf-8"))
            for k in ("cn", "overseas", "cn_model_prefixes", "overseas_model_prefixes"):
                if isinstance(ext.get(k), list):
                    raw[k] = ext[k]
        except (ValueError, OSError) as e:
            print(f"  [regions] 读取 data/regions.json 失败，使用内置默认：{e}")
    _REGION = {
        "cn": {_norm(x) for x in raw["cn"]},
        "overseas": {_norm(x) for x in raw["overseas"]},
        "cn_pre": [p.lower() for p in raw["cn_model_prefixes"]],
        "fg_pre": [p.lower() for p in raw["overseas_model_prefixes"]],
    }
    return _REGION


def _classify_region(vendor: str, model: str) -> str:
    """返回 'cn' | 'fg' | 'unknown'。组织名优先；都未命中再看模型名前缀；仍无 → unknown 并登记。"""
    reg = _load_regions()
    name, _ = _brand(vendor, model)
    for cand in (_norm(name), _norm(vendor)):
        if not cand:
            continue
        if cand in reg["cn"]:
            return "cn"
        if cand in reg["overseas"]:
            return "fg"
    m = (model or "").lower().strip()
    if any(m.startswith(p) or (" " + p) in (" " + m) for p in reg["cn_pre"]):
        return "cn"
    if any(m.startswith(p) for p in reg["fg_pre"]):
        return "fg"
    key = _norm(name) or _norm(vendor) or "(空)"
    _UNCLASSIFIED.setdefault(key, (name or vendor or "?", model))
    return "unknown"


def report_unclassified():
    """构建结束时调用：把未分类厂商写入 data/regions_unclassified.json 并打印告警；
    若本次全部已分类，则清除旧的待办文件。"""
    fp = Path(__file__).parent / "data" / "regions_unclassified.json"
    if _UNCLASSIFIED:
        items = [{"vendor": n, "example_model": mo} for n, mo in
                 (v for v in _UNCLASSIFIED.values())]
        print(f"\n  ⚠️ 发现 {len(items)} 个未分类厂商（柱色显示为灰色『未分类』）："
              f"请在 data/regions.json 的 cn / overseas 中补充：")
        for it in items:
            print(f"     · {it['vendor']}  (例: {it['example_model']})")
        try:
            fp.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"  已写入待办清单：{fp.name}")
        except OSError:
            pass
    elif fp.exists():
        try:
            fp.unlink()
        except OSError:
            pass


def _num(text):
    """提取首个数值（去千分位/%/$ 等符号），失败返回 None。"""
    if text is None:
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", str(text).replace(",", ""))
    return float(m.group(0)) if m else None


def _parse_metric_cell(text):
    """从 '11.20 % ±3.92%' 解析 (value, ci)；纯数值时 ci=None。"""
    if not text:
        return None, None
    parts = re.split(r"±", str(text))
    return _num(parts[0]), (_num(parts[1]) if len(parts) > 1 else None)


def _find(headers, *keys):
    """返回第一个表头命中任一关键词的列索引（小写包含匹配），无则 -1。"""
    for i, h in enumerate(headers):
        hl = (h or "").lower()
        if any(k in hl for k in keys):
            return i
    return -1


def _bar_cell(value, ci, axis_lo, axis_hi, max_ci, color_cls, val_disp, ci_disp, tip_rows):
    """渲染一格：① 横向柱（柱长=分值大小，按全榜分值范围缩放，区分国别配色）
    ② 右侧独立的"置信区间放大标尺"（按本榜 CI 量级缩放，把很小的 ±CI 也放大到清晰可读，
    解决基数分数大、CI 被压成一条线看不清的问题）。color_cls ∈ cn/fg/unknown。"""
    if value is None:
        return '<td class="cell-bar"><span class="bar-na">—</span></td>'

    span = (axis_hi - axis_lo) or 1.0
    fill = max(0.0, min(100.0, (value - axis_lo) / span * 100.0))

    num = f'<span class="bar-num">{_esc(val_disp)}'
    if ci_disp:
        num += f'<i class="bar-cinum">±{_esc(ci_disp)}</i>'
    num += '</span>'

    # 置信区间放大标尺：对称居中，半量程 = 本榜最大 CI（放大到清晰可读）
    if ci and max_ci:
        half = max(7.0, min(48.0, (ci / max_ci) * 48.0))   # 下限保证两端帽不重叠
        gauge = (f'<span class="ci-gauge" title="95% 置信区间 ±{_esc(ci_disp)}（已按本榜量级放大显示）">'
                 f'<span class="ci-base"></span>'
                 f'<span class="ci-span" style="left:{50 - half:.2f}%;width:{2 * half:.2f}%"></span>'
                 f'<span class="ci-est"></span></span>')
    else:
        gauge = '<span class="ci-gauge"><span class="ci-base"></span><span class="ci-est"></span></span>'

    tip = '<div class="bar-tip" hidden>'
    for label, val in tip_rows:
        if val not in (None, "", "N/A"):
            tip += (f'<div class="tip-row"><span class="tip-k">{_esc(label)}</span>'
                    f'<span class="tip-v">{_esc(str(val))}</span></div>')
    tip += '</div>'

    return (f'<td class="cell-bar" data-sort="{value}">'
            f'<div class="barc">{num}'
            f'<span class="bar-track {color_cls}">'
            f'<span class="bar-fill" style="width:{fill:.2f}%"></span></span>'
            f'{gauge}</div>{tip}</td>')


def render_category_tab(rows: list, desc: str = "") -> str:
    """渲染单个类别榜单 Tab：分值列改为横向柱状图（含置信区间须线 + 悬浮详情卡）。"""
    title = rows[0][0] if rows else ""
    headers = rows[3] if len(rows) > 3 else []
    data = rows[4:] if len(rows) > 4 else []

    html = '<div class="panel-header">'
    html += f'<h2 class="panel-title">{_esc(title)}</h2>'
    if desc:
        html += f'<p class="panel-desc">{_esc(desc)}</p>'
    html += '</div>'

    # ── 列定位（按表头）──
    i_rank = 0
    i_model = _find(headers, "模型")
    i_vendor = _find(headers, "厂商")
    i_lic = _find(headers, "许可")
    i_score = _find(headers, "elo", "分数")          # Elo 榜分值列
    i_ci = _find(headers, "置信区间")
    i_net = _find(headers, "净提升")                  # Agent 榜主指标列
    i_votes = _find(headers, "投票")
    i_spread = _find(headers, "排名区间", "spread")
    i_price = _find(headers, "价格", "price")
    i_ctx = _find(headers, "上下文", "context")
    i_sessions = _find(headers, "会话数", "session")

    is_elo = i_score >= 0
    bar_src = i_score if is_elo else i_net           # 柱状图取值列
    bar_label = "能力评分 (Elo)" if is_elo else "净提升 (Net Improvement)"

    # Agent 榜其余行为指标列（净提升之外、且非基础列）→ 进悬浮卡
    base_idx = {i_rank, i_model, i_vendor, i_lic, bar_src, i_ci}
    agent_metric_idx = [i for i, h in enumerate(headers)
                        if i not in base_idx and i != i_sessions
                        and h in METRIC_TOOLTIPS]

    # ── 预扫描：计算柱状图坐标轴范围（以最小值为基线放大差异）──
    vals = []
    for row in data:
        if is_elo:
            v = _num(row[i_score]) if i_score < len(row) else None
            c = _num(row[i_ci]) if (0 <= i_ci < len(row)) else None
        else:
            v, c = _parse_metric_cell(row[i_net]) if i_net < len(row) else (None, None)
        if v is not None:
            vals.append((v, c or 0))
    if vals:
        lo = min(v - c for v, c in vals)
        hi = max(v + c for v, c in vals)
        pad = (hi - lo) * 0.06 or 1.0
        axis_lo, axis_hi = lo - pad, hi + pad
    else:
        axis_lo, axis_hi = 0.0, 1.0
    max_ci = max((c for v, c in vals if c), default=0)   # 置信区间放大标尺的量程

    def cell(row, idx):
        return row[idx] if (0 <= idx < len(row)) else ""

    # 预计算各行国别（供柱色 + 条件图例使用）
    regions = [_classify_region(cell(row, i_vendor), cell(row, i_model)) for row in data]
    has_unknown = "unknown" in regions

    # ── 工具栏：搜索框 + 行数统计 + 国别图例（有未分类时才显示灰色项）──
    legend = ('<span class="bar-legend">'
              '<span class="lg-item"><span class="lg lg-fg"></span>国外</span>'
              '<span class="lg-item"><span class="lg lg-cn"></span>国内</span>')
    if has_unknown:
        legend += '<span class="lg-item"><span class="lg lg-unknown"></span>未分类</span>'
    legend += '</span>'
    html += ('<div class="toolbar">'
             '<input class="search-box" type="search" '
             'placeholder="🔍 搜索模型 / 厂商…" aria-label="搜索模型或厂商">'
             f'<span class="row-count" data-total="{len(data)}">共 {len(data)} 个模型</span>'
             f'{legend}</div>')

    # ── 表头 ──
    html += '<div class="table-wrapper category-table"><table>'
    bar_tip = ("柱长 = 分值高低（以最低分为基线放大差异，区分国别：国外蓝 / 国内红）；"
               "右侧小标尺 = 95% 置信区间，已按本榜量级放大显示，越宽表示越不确定；"
               "悬浮查看投票数 / 排名区间 / 价格 / 上下文。")
    if not is_elo:
        bar_tip = (METRIC_TOOLTIPS.get("净提升 (Net Improvement)", "")
                   + " 柱长 = 净提升高低；右侧小标尺 = 置信区间（已放大）；悬浮查看其余行为指标与会话数。")
    html += ('<thead><tr>'
             '<th class="col-rank">排名</th>'
             '<th class="col-model">模型名称</th>'
             '<th class="col-vendor">厂商</th>'
             '<th class="col-lic">许可</th>'
             f'<th class="col-bar">{_esc(bar_label)}'
             f'<span class="th-info" tabindex="0" title="{_esc(bar_tip)}">i</span></th>'
             '</tr></thead>')

    html += '<tbody>'
    for ri, row in enumerate(data):
        try:
            rank = int(row[i_rank]) if row[i_rank] else 0
        except (ValueError, IndexError):
            rank = 0
        rank_cls = {1: "rank-gold", 2: "rank-silver", 3: "rank-bronze"}.get(rank, "")
        medal_cls = {1: "medal-gold", 2: "medal-silver", 3: "medal-bronze"}.get(rank, "")
        model_name = cell(row, i_model)

        # 许可着色
        lic_raw = cell(row, i_lic)
        ll = lic_raw.strip().lower()
        lic_cls = ("lic-closed" if ("proprietary" in ll or "closed" in ll)
                   else "lic-open" if ("open" in ll or "apache" in ll or "mit" in ll)
                   else "lic-unknown")

        # 柱状图取值 + 展示文本
        if is_elo:
            v = _num(cell(row, i_score))
            c = _num(cell(row, i_ci))
            val_disp = (str(int(v)) if v is not None and v == int(v)
                        else (f"{v:g}" if v is not None else ""))
            ci_disp = (str(int(c)) if c is not None and c == int(c)
                       else (f"{c:g}" if c is not None else ""))
        else:
            v, c = _parse_metric_cell(cell(row, i_net))
            val_disp = f"{v:g}%" if v is not None else ""
            ci_disp = f"{c:g}%" if c is not None else ""

        # 悬浮详情卡内容
        tip_rows = [("模型", model_name)]
        if is_elo:
            tip_rows.append(("分数", f"{val_disp}{(' ± ' + ci_disp) if ci_disp else ''}"))
            tip_rows += [("投票数", cell(row, i_votes)),
                         ("排名区间", _fmt_rank_spread(cell(row, i_spread)) if i_spread >= 0 else ""),
                         ("价格", cell(row, i_price)),
                         ("上下文", cell(row, i_ctx))]
        else:
            tip_rows.append(("净提升", f"{val_disp}{(' ± ' + ci_disp) if ci_disp else ''}"))
            for mi in agent_metric_idx:
                tip_rows.append((headers[mi].split(" (")[0], cell(row, mi)))
            tip_rows.append(("会话数", cell(row, i_sessions)))

        html += f'<tr class="{rank_cls}">'
        html += (f'<td class="cell-rank {medal_cls}">{_esc(str(row[i_rank]) if i_rank < len(row) else "")}</td>'
                 f'<td class="cell-model" title="{_esc(model_name)}">{_esc(model_name)}</td>'
                 f'<td class="cell-vendor">{_vendor_cell_html(cell(row, i_vendor), model_name)}</td>'
                 f'<td class="{lic_cls}">{_esc(lic_raw)}</td>')
        html += _bar_cell(v, c, axis_lo, axis_hi, max_ci, regions[ri],
                          val_disp, ci_disp, tip_rows)
        html += '</tr>'
    html += '</tbody></table></div>'

    return html


# ── 前端交互层（普通字符串，避免 f-string 大括号转义）──
# 功能：① 表内即时搜索（模型/厂商） ② 点击表头排序（数值感知，升/降/还原三态）
#       ③ URL hash 记忆当前 Tab（可直达/可分享） ④ 过滤排序后斑马纹重排 ⑤ 回到顶部
EXTRA_CSS = """
/* ── 工具栏（搜索 + 计数） ── */
.toolbar {
    display: flex; align-items: center; gap: 12px;
    background: var(--c-surface);
    padding: 10px 14px;
    border-bottom: 1px solid var(--c-border);
}
.search-box {
    flex: 1; max-width: 320px;
    padding: 7px 12px;
    font-size: 13px;
    border: 1px solid var(--c-border);
    border-radius: 8px;
    outline: none;
    background: var(--c-bg);
    transition: border-color .15s, box-shadow .15s;
}
.search-box:focus {
    border-color: var(--c-accent);
    box-shadow: 0 0 0 3px rgba(67,97,238,.12);
    background: #fff;
}
.row-count { font-size: 12px; color: var(--c-text-secondary); white-space: nowrap; }

/* ── 可排序表头 ── */
th.sortable { cursor: pointer; user-select: none; }
th.sortable:hover { background: #233458; }
th .sort-ind { margin-left: 4px; font-size: 9px; opacity: .85; }

/* ── JS 接管后的斑马纹（过滤/排序后由 JS 重排，保持交替正确） ── */
body.js tbody tr:nth-child(even):not(.rank-gold):not(.rank-silver):not(.rank-bronze) > td {
    background: inherit;
}
body.js tr.alt:not(.rank-gold):not(.rank-silver):not(.rank-bronze) > td {
    background: #f8f9fb;
}
body.js tbody tr:hover:not(.rank-gold):not(.rank-silver):not(.rank-bronze) > td {
    background: #e8f0fe;
}

/* ── 无结果提示 ── */
.no-match {
    padding: 28px 16px; text-align: center;
    color: var(--c-text-secondary); font-size: 13px;
    background: var(--c-surface);
}

/* ── 回到顶部 ── */
.back-top {
    position: fixed; right: 18px; bottom: 22px;
    width: 40px; height: 40px;
    border: none; border-radius: 50%;
    background: var(--c-primary); color: #fff;
    font-size: 18px; cursor: pointer;
    box-shadow: var(--shadow-lg);
    opacity: 0; pointer-events: none;
    transition: opacity .2s;
    z-index: 200;
}
.back-top.show { opacity: .85; pointer-events: auto; }
.back-top:hover { opacity: 1; }
"""

SITE_JS = """
document.body.classList.add('js');

/* ── Tab 切换 + URL hash 记忆 ── */
function activateTab(tab, updateHash) {
    // 直接遍历比较 data-tab，避免 CSS 选择器对 emoji/括号的转义兼容问题
    let btn = null;
    document.querySelectorAll('.tab-btn').forEach(b => {
        if (b.dataset.tab === tab) btn = b;
    });
    const panel = document.getElementById('panel-' + tab);
    if (!btn || !panel) return false;
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
    btn.classList.add('active');
    panel.classList.add('active');
    if (updateHash) {
        try { history.replaceState(null, '', '#' + encodeURIComponent(tab)); } catch (e) {}
    }
    return true;
}
document.querySelectorAll('.tab-btn').forEach(btn => {
    btn.addEventListener('click', () => activateTab(btn.dataset.tab, true));
});
(function restoreTab() {
    const h = decodeURIComponent((location.hash || '').slice(1));
    if (h) activateTab(h, false);
})();

/* ── 通用：数值感知的单元格取值（柱状图格优先用 data-sort） ── */
function cellVal(td) {
    if (td && td.dataset && td.dataset.sort !== undefined && td.dataset.sort !== '') {
        const n = parseFloat(td.dataset.sort);
        return { n: isNaN(n) ? null : n, s: td.dataset.sort.toLowerCase() };
    }
    const t = (td ? td.textContent : '').trim();
    if (!t || t === 'N/A') return { n: null, s: t.toLowerCase() };
    const m = t.replace(/,/g, '').match(/-?\\d+(?:\\.\\d+)?/);
    return { n: m ? parseFloat(m[0]) : null, s: t.toLowerCase() };
}

/* ── 斑马纹重排（跳过前三名奖牌行与隐藏行） ── */
function restripe(tbody) {
    let i = 0;
    tbody.querySelectorAll('tr').forEach(tr => {
        tr.classList.remove('alt');
        if (tr.style.display === 'none' || tr.classList.contains('no-match-row')) return;
        const medal = tr.classList.contains('rank-gold') ||
                      tr.classList.contains('rank-silver') ||
                      tr.classList.contains('rank-bronze');
        if (!medal && i % 2 === 1) tr.classList.add('alt');
        i++;
    });
}

/* ── 表头点击排序：升 → 降 → 还原原始顺序 ── */
document.querySelectorAll('.tab-panel table').forEach(table => {
    const tbody = table.querySelector('tbody');
    if (!tbody) return;
    const rows = Array.from(tbody.querySelectorAll('tr'));
    rows.forEach((tr, i) => tr.dataset.oi = i);          // 记住原始顺序
    restripe(tbody);

    table.querySelectorAll('thead th').forEach((th, ci) => {
        th.classList.add('sortable');
        const ind = document.createElement('span');
        ind.className = 'sort-ind';
        th.appendChild(ind);

        th.addEventListener('click', () => {
            const dir = th.dataset.dir === 'asc' ? 'desc'
                      : th.dataset.dir === 'desc' ? 'none' : 'asc';
            table.querySelectorAll('thead th').forEach(o => {
                o.dataset.dir = '';
                o.querySelector('.sort-ind').textContent = '';
            });
            th.dataset.dir = dir === 'none' ? '' : dir;
            ind.textContent = dir === 'asc' ? '▲' : dir === 'desc' ? '▼' : '';

            const rs = Array.from(tbody.querySelectorAll('tr'))
                            .filter(r => !r.classList.contains('no-match-row'));
            if (dir === 'none') {
                rs.sort((a, b) => (+a.dataset.oi) - (+b.dataset.oi));
            } else {
                // 该列若多数行可解析为数值 → 数值排序，否则按文本
                const vals = rs.map(r => cellVal(r.children[ci]));
                const numeric = vals.filter(v => v.n !== null).length >= rs.length * 0.6;
                rs.map((r, i) => [r, vals[i]])
                  .sort((A, B) => {
                      const [, a] = A, [, b] = B;
                      let c;
                      if (numeric) {
                          const an = a.n === null ? Infinity : a.n;
                          const bn = b.n === null ? Infinity : b.n;
                          c = an - bn;
                      } else {
                          c = a.s.localeCompare(b.s, 'zh-Hans-CN');
                      }
                      return dir === 'asc' ? c : -c;
                  })
                  .forEach(([r]) => tbody.appendChild(r));
                restripe(tbody);
                return;
            }
            rs.forEach(r => tbody.appendChild(r));
            restripe(tbody);
        });
    });
});

/* ── 即时搜索（模型名 / 厂商，大小写不敏感；排除隐藏 tooltip 文本） ── */
function rowSearchText(tr) {
    let s = '';
    tr.querySelectorAll('td').forEach(td => {
        if (td.classList.contains('cell-bar')) {
            const num = td.querySelector('.bar-num');
            if (num) s += ' ' + num.textContent;        // 只取分值，不含隐藏卡片
        } else {
            s += ' ' + td.textContent;
        }
    });
    return s.toLowerCase();
}
document.querySelectorAll('.tab-panel').forEach(panel => {
    const box = panel.querySelector('.search-box');
    const table = panel.querySelector('table');
    const count = panel.querySelector('.row-count');
    if (!box || !table) return;
    const tbody = table.querySelector('tbody');
    const total = count ? +count.dataset.total : 0;

    let noRow = null;
    function ensureNoRow() {
        if (noRow) return noRow;
        noRow = document.createElement('tr');
        noRow.className = 'no-match-row';
        const td = document.createElement('td');
        td.colSpan = table.querySelectorAll('thead th').length || 1;
        td.className = 'no-match';
        td.textContent = '没有匹配的模型，换个关键词试试';
        noRow.appendChild(td);
        tbody.appendChild(noRow);
        return noRow;
    }

    box.addEventListener('input', () => {
        const q = box.value.trim().toLowerCase();
        let shown = 0;
        tbody.querySelectorAll('tr').forEach(tr => {
            if (tr.classList.contains('no-match-row')) return;
            const hit = !q || rowSearchText(tr).includes(q);
            tr.style.display = hit ? '' : 'none';
            if (hit) shown++;
        });
        if (count) count.textContent = q ? ('显示 ' + shown + ' / ' + total + ' 个模型')
                                         : ('共 ' + total + ' 个模型');
        ensureNoRow().style.display = shown === 0 ? '' : 'none';
        restripe(tbody);
    });
});

/* ── 柱状图悬浮详情卡（单个浮层附加到 body，避免被表格滚动容器裁剪） ── */
(function barTooltip() {
    const tip = document.createElement('div');
    tip.id = 'floating-tip';
    document.body.appendChild(tip);
    let hideT = null;

    function show(cell, ev) {
        const tpl = cell.parentElement.querySelector('.bar-tip');
        if (!tpl) return;
        clearTimeout(hideT);
        tip.innerHTML = tpl.innerHTML;
        tip.classList.add('show');
        place(ev);
    }
    function place(ev) {
        const pad = 14;
        let x = ev.clientX + pad, y = ev.clientY + pad;
        const r = tip.getBoundingClientRect();
        if (x + r.width > window.innerWidth - 8) x = ev.clientX - r.width - pad;
        if (y + r.height > window.innerHeight - 8) y = ev.clientY - r.height - pad;
        tip.style.left = Math.max(8, x) + 'px';
        tip.style.top = Math.max(8, y) + 'px';
    }
    function hide() {
        hideT = setTimeout(() => tip.classList.remove('show'), 60);
    }
    document.querySelectorAll('.cell-bar .barc').forEach(bar => {
        bar.addEventListener('mouseenter', e => show(bar, e));
        bar.addEventListener('mousemove', place);
        bar.addEventListener('mouseleave', hide);
    });
})();

/* ── 回到顶部 ── */
(function backTop() {
    const b = document.createElement('button');
    b.className = 'back-top';
    b.title = '回到顶部';
    b.textContent = '↑';
    b.addEventListener('click', () => window.scrollTo({ top: 0, behavior: 'smooth' }));
    document.body.appendChild(b);
    window.addEventListener('scroll', () => {
        b.classList.toggle('show', window.scrollY > 400);
    }, { passive: true });
})();
"""


# 「关于本榜单」卡片：仅在汇总总览页顶部展示
ABOUT_CARD_HTML = '''<details class="about-card" open>
    <summary>ℹ️ 关于本榜单 · 生成逻辑与参考价值（建议先读）</summary>
    <div class="about-body">
        <p class="about-lead">本榜单数据来自 <a href="https://arena.ai" target="_blank" rel="noopener">arena.ai</a>（前身 LMArena / Chatbot Arena，由 UC Berkeley 团队发起，累计 600&nbsp;万+ 真人投票）。它衡量的是<strong>真人偏好</strong>，而非模型的绝对能力或正确性。本项目每日自动抓取其公开数据并可视化，<strong>不做任何二次评分</strong>。</p>
        <div class="about-grid">
            <div>
                <h4>🛠️ 怎么生成的</h4>
                <p>用户在"盲测"中对两个<strong>匿名</strong>模型的回答二选一；系统用 <strong>Bradley-Terry</strong> 模型（早期为 Elo）把海量两两对比聚合成能力分数，并给出 <strong>95% 置信区间</strong>（投票越少、区间越宽）。分数为相对标尺，便于横向比较。</p>
            </div>
            <div>
                <h4>✅ 优点</h4>
                <ul>
                    <li>真人判断，而非合成题库；贴近真实使用体验</li>
                    <li>样本极大（百万级投票），匿名盲测、聚合后难以靠少量刷票操纵</li>
                    <li>覆盖文本、代码、视觉、图像/视频生成等多类别</li>
                </ul>
            </div>
            <div>
                <h4>⚠️ 局限 / 怎么看</h4>
                <ul>
                    <li><strong>"偏好"≠"能力/正确"</strong>：可能受表述风格、长度、排版讨喜程度影响</li>
                    <li>受 arena 的<strong>提问分布与投票人群</strong>影响，未必匹配你的实际场景</li>
                    <li><strong>置信区间重叠时分数小差无统计意义</strong>（约 100 分差≈64% 胜率），排名相近≈并列</li>
                    <li>新模型/小众模型投票少、区间宽，排名易波动；arena 统计口径调整也可能改变排名（参见 2025 年"Leaderboard Illusion"讨论）</li>
                </ul>
            </div>
        </div>
        <p class="about-foot">📌 建议把它当作<strong>候选筛选与趋势参考</strong>，最终选型仍应在你自己的任务与数据上实测。详细方法论见 <a href="https://arena.ai/leaderboard" target="_blank" rel="noopener">arena.ai 官方榜单</a>。</p>
    </div>
</details>'''


def build_html(sheets: dict) -> str:
    """构建完整 HTML 页面"""
    beijing_tz = timezone(timedelta(hours=8))
    now_str = datetime.now(beijing_tz).strftime("%Y-%m-%d %H:%M")

    summary_sheet = None
    category_sheets = []

    for sname, rows in sheets.items():
        if "汇总" in sname or "总览" in sname or "Overview" in sname or "Summary" in sname:
            summary_sheet = (sname, rows)
        else:
            category_sheets.append((sname, rows))

    # 硬编码 sheet 名顺序，与 export_arena_excel.py 的 CATEGORY_ORDER 一致
    SHEET_ORDER = [
        "📝 文本-对话 (Text)",
        "💻 代码 (Code)",
        "👁️ 视觉理解 (Vision)",
        "📄 文档处理 (Document)",
        "🔍 搜索增强 (Search)",
        "🤖 智能体 (Agent)",
        "🎨 文生图 (Text-to-Image)",
        "✂️ 图片编辑 (Image-Edit)",
        "🎬 文生视频 (Text-to-Video)",
        "📹 图生视频 (Image-to-Video)",
        "🎞️ 视频编辑 (Video-Edit)",
    ]
    order_map = {name: i for i, name in enumerate(SHEET_ORDER)}
    category_sheets.sort(key=lambda item: order_map.get(item[0], 99))

    # 构建 Tab 按钮和面板
    tab_buttons = []
    tab_panels = []
    active = 'active'
    if summary_sheet:
        sname, rows = summary_sheet
        tab_buttons.append(
            f'<button class="tab-btn {active}" data-tab="summary">📊 汇总总览</button>'
        )
        panel_html = render_summary_tab(category_sheets, now_str)
        tab_panels.append(f'<div class="tab-panel {active}" id="panel-summary">{ABOUT_CARD_HTML}{panel_html}</div>')
        active = ''

    for sname, rows in category_sheets:
        tab_id = sname.replace(" ", "-").replace("/", "-")
        short_name = sname.split("(")[0].strip() if "(" in sname else sname[:20]
        tab_buttons.append(
            f'<button class="tab-btn {active}" data-tab="{tab_id}">{short_name}</button>'
        )
        panel_html = render_category_tab(rows, CATEGORY_DESC.get(sname, ""))
        tab_panels.append(f'<div class="tab-panel {active}" id="panel-{tab_id}">{panel_html}</div>')
        active = ''

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="description" content="全维度 AI 模型能力排行榜：汇总 arena.ai 文本、代码、视觉、文生图、文生视频等 11 类榜单，每日自动更新，支持搜索与排序，并提供 Excel 下载。">
<meta name="theme-color" content="#1a1a2e">
<meta property="og:title" content="AI Arena Leaderboard">
<meta property="og:description" content="全维度 AI 模型能力排行榜 · 每日自动更新 · 数据来源 arena.ai">
<meta property="og:type" content="website">
<link rel="icon" href="data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'%3E%3Ctext y='.9em' font-size='90'%3E%F0%9F%8F%86%3C/text%3E%3C/svg%3E">
<title>AI Arena Leaderboard</title>
<style>
:root {{
    --c-bg: #f4f6f9;
    --c-surface: #ffffff;
    --c-primary: #1a1a2e;
    --c-primary-light: #16213e;
    --c-accent: #4361ee;
    --c-text: #1d1d1f;
    --c-text-secondary: #6e6e73;
    --c-border: #e5e5ea;
    --c-gold: #e8a817;
    --c-silver: #8e8e93;
    --c-bronze: #b5651d;
    --c-gold-bg: #fef9e7;
    --c-silver-bg: #f2f2f7;
    --c-bronze-bg: #fdf2e9;
    --c-lic-closed-fg: #0055d4;
    --c-lic-closed-bg: #e8f0fe;
    --c-lic-open-fg: #1a7f37;
    --c-lic-open-bg: #dafbe1;
    --c-lic-unknown-fg: #636366;
    --c-lic-unknown-bg: #f2f2f7;
    --radius: 10px;
    --shadow: 0 1px 3px rgba(0,0,0,.06), 0 1px 2px rgba(0,0,0,.04);
    --shadow-lg: 0 4px 12px rgba(0,0,0,.08);
}}
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC",
                 "Noto Sans SC", "Microsoft YaHei", sans-serif;
    background: var(--c-bg);
    color: var(--c-text);
    line-height: 1.5;
    -webkit-font-smoothing: antialiased;
}}

/* ── Header ── */
.header {{
    background: linear-gradient(135deg, var(--c-primary) 0%, var(--c-primary-light) 100%);
    color: #fff;
    padding: 28px 24px 20px;
    text-align: center;
}}
.header h1 {{
    font-size: 24px;
    font-weight: 700;
    letter-spacing: .5px;
    margin-bottom: 4px;
}}
.header p {{
    font-size: 13px;
    opacity: .7;
    font-weight: 400;
}}

/* ── Tab 导航 ── */
.tab-nav {{
    background: var(--c-surface);
    border-bottom: 1px solid var(--c-border);
    position: sticky;
    top: 0;
    z-index: 100;
}}
.tab-nav-inner {{
    display: flex;
    flex-wrap: wrap;
    gap: 2px 4px;
    max-width: 1200px;
    margin: 0 auto;
    padding: 4px 20px;
}}
.tab-btn {{
    flex-shrink: 0;
    border: none;
    background: transparent;
    padding: 12px 16px;
    cursor: pointer;
    font-size: 13px;
    font-weight: 500;
    color: var(--c-text-secondary);
    border-bottom: 2.5px solid transparent;
    transition: color .15s, border-color .15s;
    white-space: nowrap;
}}
.tab-btn:hover {{
    color: var(--c-text);
}}
.tab-btn.active {{
    color: var(--c-accent);
    border-bottom-color: var(--c-accent);
    font-weight: 600;
}}

/* ── 内容区 ── */
.tab-content {{
    max-width: 1200px;
    margin: 0 auto;
    padding: 16px 20px 32px;
}}
.tab-panel {{
    display: none;
}}
.tab-panel.active {{
    display: block;
}}

/* 面板标题 */
.panel-header {{
    background: var(--c-primary);
    border-radius: var(--radius) var(--radius) 0 0;
    padding: 14px 20px 12px;
}}
.panel-title {{
    color: #fff;
    font-size: 16px;
    font-weight: 600;
}}
.panel-desc {{
    color: rgba(255,255,255,.92);
    font-size: 12.5px;
    margin-top: 6px;
    line-height: 1.6;
}}
.panel-subtitle {{
    color: rgba(255,255,255,.55);
    font-size: 11px;
    margin-top: 8px;
    line-height: 1.6;
    word-break: break-all;
}}

/* ── 表格容器 ── */
.table-wrapper {{
    overflow-x: auto;
    background: var(--c-surface);
    border-radius: 0 0 var(--radius) var(--radius);
    box-shadow: var(--shadow);
    margin-bottom: 4px;
}}
table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
}}

/* ── 表头 ── */
th {{
    background: var(--c-primary-light);
    color: #fff;
    font-weight: 600;
    font-size: 12px;
    padding: 10px 12px;
    text-align: center;
    white-space: nowrap;
    position: sticky;
    top: 0;
    z-index: 1;
}}

/* 指标说明信息图标 (i) */
.th-info {{
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 14px;
    height: 14px;
    margin-left: 5px;
    border-radius: 50%;
    border: 1px solid rgba(255, 255, 255, 0.75);
    font-size: 9px;
    font-style: italic;
    font-weight: 700;
    line-height: 1;
    cursor: help;
    opacity: 0.8;
    vertical-align: middle;
    user-select: none;
}}
.th-info:hover,
.th-info:focus {{
    opacity: 1;
    background: rgba(255, 255, 255, 0.22);
    outline: none;
}}

/* ── 单元格 ── */
td {{
    padding: 9px 12px;
    border-bottom: 1px solid var(--c-border);
    text-align: center;
    vertical-align: middle;
    font-variant-numeric: tabular-nums;
}}
.cell-left {{ text-align: left; }}
.cell-model {{
    text-align: center;
    font-weight: 500;
    max-width: 280px;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
}}
.cell-rank {{
    font-weight: 700;
    width: 52px;
    font-size: 14px;
}}
.cell-score {{
    text-align: center;
    font-variant-numeric: tabular-nums;
}}
.ci-sub {{
    color: var(--c-text-secondary);
    font-size: 11px;
    font-weight: 400;
    margin-left: 3px;
}}

/* ── 厂商 logo ── */
.cell-vendor {{ text-align: center; white-space: nowrap; }}
.vlogo {{
    position: relative;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 18px; height: 18px;
    margin-right: 6px;
    border-radius: 4px;
    background: #e9ecef;
    vertical-align: middle;
    overflow: hidden;
}}
.vlogo::before {{
    content: attr(data-l);
    font-size: 10px; font-weight: 700;
    color: #6c757d;
}}
.vlogo img {{
    position: absolute;
    inset: 0;
    width: 100%; height: 100%;
    object-fit: contain;
    background: #fff;
}}
.vname {{ vertical-align: middle; }}

/* ── 排名奖牌 ── */
.medal-gold  {{ color: var(--c-gold); }}
.medal-silver {{ color: var(--c-silver); }}
.medal-bronze {{ color: var(--c-bronze); }}

tr.rank-gold > td  {{ background: var(--c-gold-bg); }}
tr.rank-silver > td {{ background: var(--c-silver-bg); }}
tr.rank-bronze > td {{ background: var(--c-bronze-bg); }}

/* ── 交替行 ── */
tbody tr:nth-child(even):not(.rank-gold):not(.rank-silver):not(.rank-bronze) > td {{
    background: #f8f9fb;
}}

/* ── Hover ── */
tbody tr:hover > td {{
    background: #e8f0fe;
}}
tbody tr.rank-gold:hover > td   {{ background: #fdf0c8; }}
tbody tr.rank-silver:hover > td {{ background: #e5e5ea; }}
tbody tr.rank-bronze:hover > td {{ background: #f9e4d1; }}

/* ── 许可文字（无色块，仅文字着色） ── */
.lic-closed, .lic-open, .lic-unknown {{
    font-size: 12px;
    font-weight: 600;
}}
.lic-closed {{ color: var(--c-lic-closed-fg); }}
.lic-open {{ color: var(--c-lic-open-fg); }}
.lic-unknown {{ color: var(--c-lic-unknown-fg); }}

/* ── 汇总表列宽 ── */
/* ── 汇总总览：各维度 Top-10 竖向柱状图 ── */
.summary-head {{
    display: flex; justify-content: space-between; align-items: flex-start; gap: 20px;
}}
.sb-updated {{
    flex: 0 0 auto; text-align: right; font-size: 11.5px;
    color: var(--c-text-secondary); line-height: 1.5; white-space: nowrap;
    display: flex; flex-direction: column;
}}
.sb-updated b {{ font-size: 15px; color: var(--c-text); font-weight: 700; margin: 1px 0; }}
.sb-legend {{
    display: flex; gap: 18px; justify-content: flex-end;
    margin: 0 2px 16px; font-size: 12px; color: var(--c-text-secondary);
}}
.sb-legend .lg-item {{ display: inline-flex; align-items: center; gap: 6px; }}

.summary-board {{
    background: var(--c-surface);
    border: 1px solid var(--c-border);
    border-radius: 12px;
    padding: 14px 18px 14px;
    margin-bottom: 14px;
    box-shadow: var(--shadow-sm, 0 1px 2px rgba(0,0,0,.04));
}}
.sb-head {{
    display: flex; justify-content: space-between; align-items: baseline;
    margin-bottom: 12px; gap: 12px;
}}
.sb-title {{ font-size: 15px; font-weight: 700; margin: 0; color: var(--c-text); }}
.sb-meta {{ font-size: 11.5px; color: var(--c-text-secondary); white-space: nowrap;
    font-variant-numeric: tabular-nums; }}
.sb-bars {{
    display: flex; align-items: flex-end; gap: 10px;
    overflow-x: auto; padding-bottom: 2px;
    position: relative;
}}
/* 数据图网格线（横向虚线）+ 基线，提升大气质感；柱体覆盖其上 */
.sb-bars::before {{
    content: ""; position: absolute; left: 0; right: 0; top: 0; height: 168px;
    background: repeating-linear-gradient(to top,
        transparent 0, transparent 41px, #edeff3 41px, #edeff3 42px);
    pointer-events: none; z-index: 0;
}}
.sb-bars::after {{
    content: ""; position: absolute; left: 0; right: 0; top: 168px; height: 2px;
    background: var(--c-border); pointer-events: none; z-index: 0;
}}
.sb-bar {{
    flex: 1 0 52px; min-width: 52px;
    display: flex; flex-direction: column; align-items: center;
    position: relative; z-index: 1;
}}
.sb-col {{ width: 100%; height: 168px; display: flex; align-items: flex-end; }}
.sb-fill {{
    width: 100%; max-width: 64px; margin: 0 auto;
    border-radius: 7px 7px 0 0;
    display: flex; justify-content: center; align-items: flex-start;
    padding-top: 9px;
    box-shadow: inset 0 -3px 8px rgba(0,0,0,.07);
    transition: filter .15s, transform .15s;
}}
.sb-bar:hover .sb-fill {{ filter: brightness(1.07) saturate(1.08); transform: translateY(-2px); }}
.sb-score {{
    color: #fff; font-weight: 700; font-size: 13px;
    font-variant-numeric: tabular-nums;
    text-shadow: 0 1px 2px rgba(0,0,0,.28);
}}
.sb-logo {{ margin-top: 9px; }}
.sb-logo .vlogo {{ width: 24px; height: 24px; margin-right: 0; border-radius: 5px; }}
.sb-logo .vlogo::before {{ font-size: 12px; }}
.sb-name {{
    margin-top: 6px; font-size: 10.5px; line-height: 1.25; text-align: center;
    color: var(--c-text-secondary);
    display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical;
    overflow: hidden; height: 2.5em; word-break: break-word; width: 100%;
}}
/* 柱色按国别（与各榜单一致） */
.sb-bar.fg .sb-fill {{ background: linear-gradient(180deg,#6f9bf2,#3f6fe0); }}
.sb-bar.cn .sb-fill {{ background: linear-gradient(180deg,#f47b7d,#e23b3e); }}
.sb-bar.unknown .sb-fill {{ background: linear-gradient(180deg,#cdd2d9,#9aa1ab); }}

/* ── 类别表列宽（精简为：排名/模型/厂商/许可/柱状图）── */
.category-table th.col-rank,
.category-table td.cell-rank {{ width: 52px; }}
.category-table th.col-model,
.category-table td.cell-model {{ min-width: 180px; }}
.category-table th.col-vendor,
.category-table td.cell-vendor {{ min-width: 110px; }}
.category-table th.col-lic,
.category-table td:nth-child(4) {{ width: 96px; }}
.category-table th.col-bar {{ min-width: 340px; }}

/* ── 横向柱状图单元格 ── */
.cell-bar {{ text-align: left; padding: 8px 16px 8px 10px; }}
.barc {{
    display: flex;
    align-items: center;
    gap: 10px;
    min-width: 240px;
}}
.bar-num {{
    flex: 0 0 auto;
    width: 82px;                 /* 固定宽度 + 左对齐：数值左边、色条左边各自跨行对齐 */
    text-align: left;
    white-space: nowrap;
    font-weight: 700;
    font-size: 13px;
    font-variant-numeric: tabular-nums;
    color: var(--c-text);
}}
.bar-cinum {{
    color: var(--c-text-secondary);
    font-size: 10.5px;
    font-weight: 500;
    font-style: normal;
    margin-left: 2px;
}}
.bar-track {{
    position: relative;
    flex: 1 1 auto;
    height: 22px;
    background: #eef0f3;
    border-radius: 5px;
    overflow: hidden;
}}
.bar-fill {{
    position: absolute;
    left: 0; top: 0; bottom: 0;
    border-radius: 5px;
    min-width: 2px;
    transition: filter .15s;
}}
.bar-na {{ color: var(--c-text-secondary); }}

/* ── 置信区间「放大标尺」：按本榜 CI 量级缩放，对称居中，永远清晰可读 ── */
.ci-gauge {{
    position: relative;
    flex: 0 0 auto;
    width: 76px;
    height: 22px;
    margin-left: 2px;
}}
.ci-base {{                       /* 标尺基线 */
    position: absolute;
    left: 4px; right: 4px; top: 50%;
    height: 0;
    border-top: 1px dashed #c7ccd4;
    transform: translateY(-50%);
}}
.ci-span {{                       /* 置信区间段（带两端竖帽） */
    position: absolute;
    top: 50%;
    height: 0;
    border-top: 2px solid rgba(60,60,67,.62);
    transform: translateY(-50%);
}}
.ci-span::before, .ci-span::after {{
    content: "";
    position: absolute;
    top: 50%;
    width: 0; height: 11px;
    border-left: 2px solid rgba(60,60,67,.62);
    transform: translateY(-50%);
}}
.ci-span::before {{ left: 0; }}
.ci-span::after {{ right: 0; }}
.ci-est {{                        /* 中心估计点 */
    position: absolute;
    left: 50%; top: 50%;
    width: 7px; height: 7px;
    border-radius: 50%;
    background: #2c2c3a;
    transform: translate(-50%, -50%);
    z-index: 2;
}}

/* 柱色按国别：国外蓝 / 国内红 / 未分类灰（名次由整行背景色区分） */
.bar-track.fg .bar-fill {{ background: linear-gradient(90deg,#5b8def,#3f6fe0); }}
.bar-track.cn .bar-fill {{ background: linear-gradient(90deg,#f0696b,#e23b3e); }}
.bar-track.unknown .bar-fill {{ background: linear-gradient(90deg,#c2c7cf,#9aa1ab); }}
tbody tr:hover .bar-fill {{ filter: brightness(1.06) saturate(1.08); }}

/* 国别图例（工具栏右侧） */
.bar-legend {{
    margin-left: auto;
    display: inline-flex;
    align-items: center;
    gap: 12px;
    font-size: 12px;
    color: var(--c-text-secondary);
    white-space: nowrap;
}}
.bar-legend .lg-item {{ display: inline-flex; align-items: center; gap: 5px; }}
.bar-legend .lg {{
    width: 18px; height: 10px; border-radius: 3px; display: inline-block;
}}
.bar-legend .lg-fg {{ background: linear-gradient(90deg,#5b8def,#3f6fe0); }}
.bar-legend .lg-cn {{ background: linear-gradient(90deg,#f0696b,#e23b3e); }}
.bar-legend .lg-unknown {{ background: linear-gradient(90deg,#c2c7cf,#9aa1ab); }}

/* ── 悬浮详情卡（JS 接管定位，附加到 body，防裁剪）── */
.bar-tip {{ display: none; }}
#floating-tip {{
    position: fixed;
    z-index: 999;
    min-width: 180px;
    max-width: 280px;
    background: #1f2433;
    color: #f2f4f8;
    border-radius: 9px;
    padding: 10px 12px;
    font-size: 12px;
    line-height: 1.55;
    box-shadow: 0 6px 22px rgba(0,0,0,.28);
    pointer-events: none;
    opacity: 0;
    transition: opacity .12s;
}}
#floating-tip.show {{ opacity: 1; }}
#floating-tip .tip-row {{
    display: flex; justify-content: space-between; gap: 16px;
    padding: 2px 0;
}}
#floating-tip .tip-row:first-child {{
    margin-bottom: 4px; padding-bottom: 6px;
    border-bottom: 1px solid rgba(255,255,255,.14);
}}
#floating-tip .tip-row:first-child .tip-k {{ display: none; }}
#floating-tip .tip-row:first-child .tip-v {{
    font-weight: 700; font-size: 13px; color: #fff;
}}
#floating-tip .tip-k {{ color: rgba(255,255,255,.6); white-space: nowrap; }}
#floating-tip .tip-v {{ font-weight: 600; font-variant-numeric: tabular-nums; }}

/* ── Footer ── */
.footer {{
    text-align: center;
    color: var(--c-text-secondary);
    font-size: 12px;
    padding: 16px 24px 32px;
    opacity: .6;
}}

/* ── 响应式 ── */
@media (max-width: 768px) {{
    .header {{ padding: 20px 16px 16px; }}
    .header h1 {{ font-size: 20px; }}
    .tab-nav-inner {{ padding: 0 12px; }}
    .tab-btn {{ padding: 10px 12px; font-size: 12px; }}
    .tab-content {{ padding: 12px; }}
    table {{ font-size: 12px; }}
    th {{ padding: 8px 6px; font-size: 11px; }}
    td {{ padding: 7px 6px; }}
    .cell-model {{ max-width: 160px; }}
    .panel-header {{ padding: 12px 14px 10px; }}
    .panel-title {{ font-size: 14px; }}
}}
/* 关于本榜单卡片 */
.about-card {{
    margin: 0 0 16px;
    background: var(--c-surface, #fff);
    border: 1px solid var(--c-border, #e5e7eb);
    border-left: 4px solid #2b6cb0;
    border-radius: 10px;
    box-shadow: 0 1px 3px rgba(0,0,0,.06);
    overflow: hidden;
}}
.about-card > summary {{
    cursor: pointer;
    list-style: none;
    padding: 12px 16px;
    font-weight: 700;
    font-size: 14px;
    color: #1a2b4a;
    background: #f3f7fc;
    user-select: none;
}}
.about-card > summary::-webkit-details-marker {{ display: none; }}
.about-card > summary::after {{ content: " ▸"; color: #2b6cb0; }}
.about-card[open] > summary::after {{ content: " ▾"; }}
.about-body {{ padding: 4px 18px 16px; }}
.about-lead {{ font-size: 13px; line-height: 1.7; color: #374151; margin: 10px 0 14px; }}
.about-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
    gap: 16px 24px;
}}
.about-grid h4 {{ font-size: 13px; margin: 0 0 6px; color: #1a2b4a; }}
.about-grid p, .about-grid li {{ font-size: 12.5px; line-height: 1.65; color: #4b5563; }}
.about-grid ul {{ margin: 0; padding-left: 18px; }}
.about-grid li {{ margin-bottom: 4px; }}
.about-foot {{
    font-size: 12.5px; line-height: 1.65; color: #374151;
    margin: 14px 0 0; padding-top: 12px; border-top: 1px solid var(--c-border, #eee);
}}
.about-card a {{ color: #2b6cb0; }}
@media (max-width: 600px) {{
    .about-grid {{ grid-template-columns: 1fr; gap: 12px; }}
}}
/* 厂商图标（构建时内嵌，每图标仅定义一次） */
{logo_css()}
/* 前端交互层（搜索 / 排序 / Tab 记忆 / 回到顶部） */
{EXTRA_CSS}
</style>
</head>
<body>
<div class="header">
    <h1>AI Arena Leaderboard</h1>
    <p>全维度 AI 模型能力排行榜 · 数据来源 arena.ai · 更新时间 {now_str} (北京时间)</p>
</div>

<nav class="tab-nav">
    <div class="tab-nav-inner">{"".join(tab_buttons)}</div>
</nav>

<div class="tab-content">
    {"".join(tab_panels)}
</div>

<div class="footer">
    <p>数据每日定时自动更新 · Powered by GitHub Actions</p>
</div>

<script>
{SITE_JS}
</script>
</body>
</html>'''
    return html


def main():
    global _LOGO_REFRESH
    _LOGO_REFRESH = "--refresh-logos" in sys.argv   # 强制忽略缓存、重新直抓官网图标

    xlsx_dir = Path(__file__).parent
    xlsx_path = xlsx_dir / "arena_leaderboard.xlsx"

    if not xlsx_path.exists():
        candidates = sorted(xlsx_dir.glob("arena_leaderboard*.xlsx"), reverse=True)
        if not candidates:
            print("No xlsx file found. Run export_arena_excel.py first.")
            sys.exit(1)
        xlsx_path = candidates[0]

    print(f"Reading: {xlsx_path}")
    sheets = extract_sheets(str(xlsx_path))
    load_logo_cache()
    print(f"Resolving vendor logos (cached: {sum(1 for v in _RESOLVED_LOGOS.values() if v)})…")
    html = build_html(sheets)
    save_logo_cache()
    report_unclassified()
    out_path = xlsx_dir / "index.html"
    out_path.write_text(html, encoding="utf-8")
    print(f"Written: {out_path}")


if __name__ == "__main__":
    main()