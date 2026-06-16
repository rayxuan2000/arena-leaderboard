#!/usr/bin/env python3
"""
AI Arena Leaderboard Excel Exporter

将 AI Arena 各类别排行榜 JSON 数据聚合导出为一个格式美观的 Excel 文件，
每个类别一个 Sheet，包含条件格式、排名可视化、汇总分析等功能。

用法:
    python export_arena_excel.py [--input-dir <json目录>] [--output <输出excel路径>]
"""

import sys
import json
import argparse
from pathlib import Path
from datetime import datetime, timezone, timedelta

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter

# ── 配色方案 ──────────────────────────────────────────────
COLORS = {
    "primary":    "1a1a2e",   # 深蓝黑（标题背景）
    "secondary":  "16213e",   # 深蓝（表头背景）
    "accent":     "e94560",   # 红粉（强调色）
    "gold":       "f0a500",   # 金色（排名第一）
    "silver":     "c0c0c0",   # 银色（排名第二）
    "bronze":     "cd7f32",   # 铜色（排名第三）
    "light_bg":   "f8f9fa",   # 浅灰背景
    "white":      "ffffff",   # 白色
    "text_dark":  "212529",   # 深色文字
    "text_light": "ffffff",   # 浅色文字
    "border":     "dee2e6",   # 边框色
    "green":      "28a745",   # 绿色（open license）
    "blue":       "007bff",   # 蓝色（proprietary）
    "orange":     "fd7e14",   # 橙色
}

# ── 样式定义 ──────────────────────────────────────────────
thin_border = Border(
    left=Side(style="thin", color=COLORS["border"]),
    right=Side(style="thin", color=COLORS["border"]),
    top=Side(style="thin", color=COLORS["border"]),
    bottom=Side(style="thin", color=COLORS["border"]),
)

header_fill = PatternFill(start_color=COLORS["secondary"], end_color=COLORS["secondary"], fill_type="solid")
header_font = Font(name="Arial", bold=True, color=COLORS["text_light"], size=11)
title_fill = PatternFill(start_color=COLORS["primary"], end_color=COLORS["primary"], fill_type="solid")
title_font = Font(name="Arial", bold=True, color=COLORS["text_light"], size=14)
subtitle_font = Font(name="Arial", color=COLORS["text_light"], size=10)
data_font = Font(name="Arial", size=11, color=COLORS["text_dark"])
alt_row_fill = PatternFill(start_color="edf2f9", end_color="edf2f9", fill_type="solid")
gold_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
silver_fill = PatternFill(start_color="e2e3e5", end_color="e2e3e5", fill_type="solid")
bronze_fill = PatternFill(start_color="fde8d0", end_color="fde8d0", fill_type="solid")

# ── 类别显示名称映射 ──────────────────────────────────────
CATEGORY_NAMES = {
    "text":             "📝 文本/对话 (Text)",
    "code":             "💻 代码 (Code)",
    "vision":           "👁️ 视觉理解 (Vision)",
    "document":         "📄 文档处理 (Document)",
    "search":           "🔍 搜索增强 (Search)",
    "agent":            "🤖 智能体 (Agent)",
    "text-to-image":    "🎨 文生图 (Text-to-Image)",
    "image-edit":       "✂️ 图片编辑 (Image-Edit)",
    "text-to-video":    "🎬 文生视频 (Text-to-Video)",
    "image-to-video":   "📹 图生视频 (Image-to-Video)",
    "video-edit":       "🎞️ 视频编辑 (Video-Edit)",
}

CATEGORY_ORDER = [
    "text", "code", "vision", "document", "search", "agent",
    "text-to-image", "image-edit",
    "text-to-video", "image-to-video", "video-edit",
]

LICENSE_COLORS = {
    "proprietary": ("007bff", "cfe2ff"),   # 蓝
    "open":        ("28a745", "d1e7dd"),   # 绿
    None:          ("6c757d", "e9ecef"),   # 灰
}


def fetch_leaderboard_data(data_dir: str = "data") -> dict:
    """
    读取由 fetch_arena.py 直连 arena.ai 抓取并落地的榜单 JSON。
    优先读取合并文件 data/arena_data.json，其次逐个读取 data/<slug>.json。
    若本地没有数据，则就地调用 fetch_arena 直接抓取（保证单独运行也可用）。
    """
    ddir = Path(data_dir)
    data = {}

    combined = ddir / "arena_data.json"
    if combined.exists():
        loaded = json.loads(combined.read_text(encoding="utf-8"))
        for category in CATEGORY_ORDER:
            if category in loaded and loaded[category].get("models"):
                data[category] = loaded[category]
                print(f"[OK] Loaded: {category} ({loaded[category]['meta']['model_count']})")
    else:
        for category in CATEGORY_ORDER:
            fp = ddir / f"{category}.json"
            if fp.exists():
                obj = json.loads(fp.read_text(encoding="utf-8"))
                if obj.get("models"):
                    data[category] = obj
                    print(f"[OK] Loaded: {category} ({obj['meta']['model_count']})")

    # 本地无数据 → 就地直连 arena.ai 抓取
    if not data:
        print("[INFO] No local data found, fetching directly from arena.ai...")
        try:
            sys.path.insert(0, str(Path(__file__).resolve().parent))
            import fetch_arena
            for category in CATEGORY_ORDER:
                res = fetch_arena.fetch_one(category, allow_jina=False)
                if res["meta"]["model_count"] > 0:
                    data[category] = res
                    print(f"[OK] Fetched: {category} ({res['meta']['model_count']})")
                else:
                    print(f"[WARN] No data for {category}")
        except Exception as e:  # noqa: BLE001
            print(f"[ERROR] Direct fetch failed: {e}")

    return data


def create_summary_sheet(wb: Workbook, data: dict):
    """创建汇总 Sheet"""
    ws = wb.active
    ws.title = "📊 汇总总览"

    # 标题行
    ws.merge_cells("A1:H1")
    ws["A1"] = "AI Arena 排行榜总览"
    ws["A1"].font = Font(name="Arial", bold=True, color=COLORS["text_light"], size=18)
    ws["A1"].fill = PatternFill(start_color=COLORS["primary"], end_color=COLORS["primary"], fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42

    # 副标题
    ws.merge_cells("A2:H2")
    # 与网站页脚保持一致，统一用北京时间（CI runner 默认时钟是 UTC，
    # 若用 naive datetime.now() 会比页面晚 8 小时，造成 Excel 与网站时间对不上）。
    _bj_now = datetime.now(timezone(timedelta(hours=8))).strftime('%Y-%m-%d %H:%M')
    ws["A2"] = f"数据来源: arena.ai  |  生成时间: {_bj_now} (北京时间)"
    ws["A2"].font = Font(name="Arial", color=COLORS["text_light"], size=10, italic=True)
    ws["A2"].fill = PatternFill(start_color=COLORS["primary"], end_color=COLORS["primary"], fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # 空行
    ws.row_dimensions[3].height = 12

    # 表头
    headers = ["类别", "模型数", "第1名", "第1名厂商", "第1名分数", "第2名", "第3名", "数据更新日期"]
    col_widths = [28, 10, 28, 16, 14, 28, 28, 18]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 28

    # 数据行
    row = 5
    for category in CATEGORY_ORDER:
        if category not in data:
            continue
        meta = data[category].get("meta", {})
        models = data[category].get("models", [])

        display_name = CATEGORY_NAMES.get(category, category)

        values = [
            display_name,
            meta.get("model_count", len(models)),
            models[0].get("model", "") if len(models) > 0 else "",
            models[0].get("vendor", "") if len(models) > 0 else "",
            models[0].get("score", "") if len(models) > 0 else "",
            models[1].get("model", "") if len(models) > 1 else "",
            models[2].get("model", "") if len(models) > 2 else "",
            meta.get("last_updated") or "N/A",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center" if col_idx != 1 else "left", vertical="center")
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = alt_row_fill

        ws.row_dimensions[row].height = 26
        row += 1

    # 冻结窗格
    ws.freeze_panes = "A5"


# Agent 榜行为指标列 → 中文显示名（其余未知键回退为 Title Case）
AGENT_METRIC_LABELS = {
    "net improvement": "净提升 (Net Improvement)",
    "confirmed success": "确认完成 (Confirmed Success)",
    "praise vs complaint": "好评/差评 (Praise vs Complaint)",
    "steerability": "可操控性 (Steerability)",
    "bash recovery": "命令恢复 (Bash Recovery)",
    "tool hallucination": "工具幻觉 (Tool Hallucination)",
    "sessions": "会话数 (Sessions)",
}


# 非核心列（rank spread / price / context 等）→ 中文显示名；未知键回退 Title Case
def col_label(key: str) -> str:
    k = (key or "").strip().lower()
    if k in AGENT_METRIC_LABELS:
        return AGENT_METRIC_LABELS[k]
    if "price" in k or "cost" in k or "$/m" in k:
        return "价格 ($/M)"
    if "context" in k:
        return "上下文长度"
    if "spread" in k:
        return "排名区间"
    return key.title() if key else ""


def create_leaderboard_sheet(wb: Workbook, category: str, data: dict):
    """为单个类别创建排行榜 Sheet（Elo 榜与 Agent 指标榜列结构不同，自动适配）"""
    meta = data.get("meta", {})
    models = data.get("models", [])

    display_name = CATEGORY_NAMES.get(category, category)
    # Excel sheet name: max 31 chars, no \ / * ? : [ ]
    sheet_name = display_name
    for ch in "\\/*?:[]":
        sheet_name = sheet_name.replace(ch, "-")
    sheet_name = sheet_name[:31]

    ws = wb.create_sheet(title=sheet_name)

    # ── 列布局：Elo 榜（有分数）vs 指标榜（Agent，无 Elo/票数，列为行为指标）──
    is_metric = bool(models) and not any(m.get("score") for m in models)
    if is_metric:
        metric_keys = []
        for m in models:
            for k in (m.get("extra") or {}):
                if k not in metric_keys:
                    metric_keys.append(k)
        headers = (["排名", "模型名称", "厂商", "许可"]
                   + [col_label(k) for k in metric_keys]
                   + ["会话数 (Sessions)"])
        col_widths = [8, 34, 16, 13] + [17] * len(metric_keys) + [13]
    else:
        # Elo 榜：核心 7 列之外，把 extra（rank spread / price / context 等）按原序铺开，1:1 复刻
        extra_keys = []
        for m in models:
            for k in (m.get("extra") or {}):
                if k not in extra_keys:
                    extra_keys.append(k)
        headers = (["排名", "模型名称", "厂商", "许可", "Elo 分数", "置信区间(±)", "投票数"]
                   + [col_label(k) for k in extra_keys])
        col_widths = [8, 38, 18, 14, 14, 16, 14] + [16] * len(extra_keys)
    ncols = len(headers)
    last_col = get_column_letter(ncols)

    # ── 标题区域 ──
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"AI Arena 排行榜 - {display_name}"
    ws["A1"].font = Font(name="Arial", bold=True, color=COLORS["text_light"], size=16)
    ws["A1"].fill = PatternFill(start_color=COLORS["primary"], end_color=COLORS["primary"], fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 第 2 行保留为空白占位行（不再写入来源/更新/抓取时间等元数据），
    # 以维持 build_site.py 对表头(第4行)和数据(第5行起)行号的依赖。
    ws.row_dimensions[2].height = 6

    # 空行
    ws.row_dimensions[3].height = 8

    # ── 表头 ──
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 30

    # ── 数据行 ──
    for i, model in enumerate(models):
        row = 5 + i
        rank = model.get("rank", i + 1)
        license_type = model.get("license")

        base = [rank, model.get("model", ""), model.get("vendor", ""), license_type or "N/A"]
        if is_metric:
            ex = model.get("extra") or {}
            values = base + [ex.get(k, "") for k in metric_keys] + [model.get("votes", "")]
        else:
            ex = model.get("extra") or {}
            values = (base
                      + [model.get("score", ""), model.get("ci", ""), model.get("votes", "")]
                      + [ex.get(k, "") for k in extra_keys])

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left", vertical="center")
            cell.border = thin_border

        # 交替行颜色
        if i % 2 == 1:
            for col_idx in range(1, ncols + 1):
                ws.cell(row=row, column=col_idx).fill = alt_row_fill

        # 排名前三高亮
        if rank == 1:
            for col_idx in range(1, ncols + 1):
                ws.cell(row=row, column=col_idx).fill = gold_fill
            ws.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=13, color=COLORS["gold"])
        elif rank == 2:
            ws.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=12, color="808080")
        elif rank == 3:
            ws.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=12, color=COLORS["bronze"])

        # 许可类型颜色标记
        lic_cell = ws.cell(row=row, column=4)
        lic_color = LICENSE_COLORS.get(license_type, LICENSE_COLORS[None])
        lic_cell.font = Font(name="Arial", size=10, bold=True, color=lic_color[0])
        lic_cell.fill = PatternFill(start_color=lic_color[1], end_color=lic_color[1], fill_type="solid")

        # 分数格式（仅 Elo 榜）
        if not is_metric:
            ws.cell(row=row, column=5).number_format = "#,##0"

        ws.row_dimensions[row].height = 24

    last_data_row = 4 + len(models)

    # ── 分数列条件格式（仅 Elo 榜；指标榜的值是带 ± 的百分比字符串，不适用）──
    if not is_metric and models:
        ws.conditional_formatting.add(
            f"E5:E{last_data_row}",
            DataBarRule(
                start_type="min", end_type="max",
                color=COLORS["accent"],
                showValue=True,
                minLength=None, maxLength=None,
            ),
        )
        ws.conditional_formatting.add(
            f"E5:E{last_data_row}",
            ColorScaleRule(
                start_type="min", start_color="f8d7da",
                mid_type="percentile", mid_value=50, mid_color="fff3cd",
                end_type="max", end_color="d1e7dd",
            ),
        )

    # ── 冻结表头 ──
    ws.freeze_panes = "A5"

    # ── 自动筛选 ──
    ws.auto_filter.ref = f"A4:{last_col}{last_data_row}"


def main():
    parser = argparse.ArgumentParser(description="AI Arena 排行榜 Excel 导出工具")
    parser.add_argument(
        "--output",
        default=None,
        help="输出 Excel 文件路径 (默认: 当前目录下的 arena_leaderboard.xlsx)",
    )
    parser.add_argument(
        "--input-dir",
        default="data",
        help="fetch_arena.py 落地的 JSON 目录 (默认: ./data)",
    )
    args = parser.parse_args()

    # 确定输出路径
    if args.output:
        output_path = args.output
    else:
        output_path = str(Path.cwd() / "arena_leaderboard.xlsx")

    # 读取 fetch_arena.py 直连 arena.ai 落地的本地 JSON（无则就地直抓兜底）
    print(f"[FETCH] Loading leaderboard data from '{args.input_dir}/' ...")
    data = fetch_leaderboard_data(args.input_dir)

    if not data:
        print("[ERROR] Failed to fetch any leaderboard data.")
        sys.exit(1)

    print(f"[OK] Fetched {len(data)} leaderboard categories")

    # 创建工作簿
    wb = Workbook()

    # 创建汇总 Sheet
    print("[INFO] Creating summary sheet...")
    create_summary_sheet(wb, data)

    # 为每个类别创建 Sheet
    for category in CATEGORY_ORDER:
        if category in data:
            print(f"[INFO] Creating sheet: {category}")
            create_leaderboard_sheet(wb, category, data[category])

    # 保存
    wb.save(output_path)
    print(f"\n[DONE] Export completed: {output_path}")
    print(f"       Total: {len(data)} leaderboard sheets + 1 summary sheet")


if __name__ == "__main__":
    main()